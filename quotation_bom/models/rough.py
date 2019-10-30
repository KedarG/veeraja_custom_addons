# import formulas
# import xlrd
import openpyxl
wb = openpyxl.load_workbook("/home/jeevan/Music/easy.xlsx")
sh1 = wb.get_sheet_by_name('15')


# i = 0
# for x in range(4000, 5014):
#     temp = sh1[x][2].value
#     if temp != '' or temp != '=':
#         a = str(temp).split('/')
#         sep = ' / '
#         sh1[x][2].value = sep.join(a)
#         i += 1
#         print("count", i)
# wb.save('/home/jeevan/veeraja/trial_Veeraja_Item_Master.xlsx')
# print(sh1[x][2].value)
# sh1[]
sh1['M9'].value = 4000000
wb.save('/home/jeevan/Pictures/new_105tst.xlsx')
wb.close()

new_wb = openpyxl.load_workbook(
    "/home/jeevan/Pictures/new_105tst.xlsx", data_only=True)
new_sh = new_wb.get_sheet_by_name('15')
print("M9====>", new_sh['M9'].value)
print("J9====>", new_sh['J9'].value)
# xlwb1 = xlrd.open_workbook(filename='/home/jeevan/Pictures/test_102.xlsx')
# xlsh1 = xlwb1.sheet_by_name('15')
# print("xlsh1.cell(8,9)", xlsh1.cell(8, 9), xlsh1.cell(8, 12))
# fpath = '/home/jeevan/new_odoo-11.0/easy.xlsx'
# wb = formulas.ExcelModel().loads(fpath).finish()
# opwb = openpyxl.load_workbook(fpath)
# wb.calculate()
# sh1 = opwb.get_sheet_by_name('15')
# print('===============>', sh1['J9'].value)
# from win32com.client import Dispatch
# excel = Dispatch("Excel.Application")
# wb = excel.Workbooks.Append('/home/jeevan/Pictures/easy_output_104.xlsx')
# range1 = wb.Sheets[0].Range("J9")
# val=range1.Value


{
    'name': '',
    'version': '11.0.1.0.0',
    'summary': '',
    'category': '',
    'author': '',
    'maintainer': '',
    'website': '',
    'license': '',
    'contributors': [
        '',
    ],
    'depends': [
        '',
    ],
    'external_dependencies': {
        'python': [
            '',
        ],
    },
    'data': [
        '',
    ],
    'installable': True,
    'auto_install': False,
    'application': False,
}
