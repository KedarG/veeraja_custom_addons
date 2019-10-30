from odoo import models, fields, api, _
import openpyxl
import xlrd
import base64
import codecs
from odoo.exceptions import UserError
from io import BytesIO, StringIO


class TransactionInfo(models.Model):
    _name = 'transaction.info'
    _rec_name = 'name'
    @api.model
    def set_formula(self):
        if 'default_product_category_id' in self.env.context:
            w = self.env.context['default_product_category_id']
            li = []
            domain = {}
            formula_info_obj = self.env['formula.info'].search(
                [('product_category_id', '=', w)])
            if formula_info_obj:
                for formula in formula_info_obj:
                    li.append(formula.id)
                domain['select_formula_name'] = [('id', 'in', li)]
                return [('id', 'in', li)]
    name = fields.Char(
        string='Name', default=lambda self: _('New'), store=True)
    select_formula_name = fields.Many2one(
        "formula.info", string='Formula', domain=lambda self: self.set_formula(), store=True, readonly=False)
    transaction_line_id = fields.One2many(
        "transaction.details", 'transaction_id')
    product_ids = fields.One2many(
        "transaction.details.products", 'transaction_id')
    total = fields.Float("Total Amount", compute="set_total", store=True)
    product_category_id = fields.Many2one(
        "product.category", string="Product Category")
    qbom_line_id = fields.Many2one(
        'quotation.bom.line', string="Quotation Bom Line Id")
    excel_formula = fields.Binary(string="Excel Formula")
    output_excel = fields.Binary(string="Excel With Calulated Values")
    filename = fields.Char()
    new_file = fields.Char()
    excel_sheet_name = fields.Char(string="Excel Sheet Name")
    total_weight = fields.Float(
        string="Total Weight", store=True, compute='set_total_weight')

    @api.depends('product_ids.weight')
    def set_total_weight(self):
        for order in self:
            if order.product_ids:
                weight_tot = 0
                for i in order.product_ids:
                    weight_tot += i.weight or 0.0
                order.total_weight = weight_tot

    @api.multi
    def set_prods_excel(self):
        if self.output_excel:
            prod_obj = self.env['product.product']
            ax1 = base64.b64decode(codecs.decode(self.output_excel))
            book = xlrd.open_workbook(file_contents=ax1)
            sheet_names = book.sheet_names()
            index_of_sheet = sheet_names.index(self.excel_sheet_name)
            first_sheet = book.sheet_by_name(sheet_names[index_of_sheet])
            ro = first_sheet.nrows
            co = first_sheet.ncols
            data = []
            for i in range(39, ro):
                maja = []
                asdf = 0
                for j in range(0, co):
                    cell = first_sheet.cell(i, j)
                    maja.append(cell.value)
                    asdf += 1
                data.append(maja)
            prods = []
            for rec in data:
                functional_categ = rec[0]
                item_code = rec[4]
                description = rec[7]
                qty = rec[17]
                if qty != 0.0:
                    if item_code:
                        prod_id = prod_obj.search(
                            [('default_code', '=', item_code)])
                        functional_categ_id = self.env['product.category'].search(
                            [('name', '=', functional_categ)])
                        if prod_id:
                            values = {
                                'product_id': prod_id.id,
                                'cost': prod_id.standard_price,
                                'description': description,
                                'product_qty': qty,
                                'functional_categ_id': functional_categ_id.id if functional_categ_id else ''
                            }
                            prods.append((0, 0, values))
                        else:
                            values = {
                                'description': description,
                                'product_qty': qty,
                                'functional_categ_id': functional_categ_id.id if functional_categ_id else ''
                            }
                            prods.append((0, 0, values))
                    else:
                        raise UserError(_('Item Code Not Present'))
            self.product_ids.unlink()
            self.product_ids = [prod for prod in prods]

    @api.depends('product_ids.cost')
    def set_total(self):
        for transaction in self:
            if transaction.product_ids:
                amount = 0.00
                for t_line in transaction.product_ids:
                    if t_line.cost:
                        amount += t_line.cost
                        # print(amount)
                    transaction.total = amount

    # def context_values(self):
    #     formula_detail = self.env['formula.details'].search(
    #         [('formula_id', '=', self.select_formula_name.id)])
    #     li = []
    #     for order in formula_detail:
    #         values = {
    #             "trans_process": order.process,
    #             'trans_name': order.name,
    #             'trans_parameter': order.parameter_name,
    #             'trans_uom': order.uom,
    #             'trans_formula_expression': order.formula_expression,
    #             'detail_product_category_id': order.detail_product_category_id
    #         }
    #         li.append((0, 0, values))
    #     return li

    @api.onchange('select_formula_name')
    def get_record_db(self):
        if self.select_formula_name:
            self.filename = self.select_formula_name.filename
            self.excel_sheet_name = self.select_formula_name.excel_sheet_name
            self.excel_formula = self.select_formula_name.excel_formula
            # self.transaction_line_id = [j for j in self.context_values()]

    # Discarded function
    def set_old_products(self):
        # for transaction in self:
        #     if transaction.transaction_line_id:
        #         li = []
        #         for t_line in transaction.transaction_line_id:
        #             if t_line.detail_product_category_id:
        #                 prod_temp_ids = self.env['product.template'].search(
        #                     [('option', '=', t_line.usr_input), ('categ_id', '=', t_line.detail_product_category_id.id)])
        #                 for j in prod_temp_ids:
        #                     prod_ids = self.env['product.product'].search(
        #                         [('product_tmpl_id', '=', j.id)])
        #                     if prod_ids:
        #                         for i in prod_ids:
        #                             values = {
        #                                 'product_id': i.id,
        #                                 'cost': i.standard_price
        #                             }
        #                             transaction.qbom_line_id.unit_price = values['cost']
        #                             li.append((0, 0, values))
        #                 if not prod_temp_ids:
        #                     prod_prod_ids = self.env['product.product'].search(
        #                         [('option', '=', t_line.usr_input), ('categ_id', '=', t_line.detail_product_category_id.id)])
        #                     if prod_prod_ids:
        #                         for z in prod_prod_ids:
        #                             values = {
        #                                 'product_id': z.id,
        #                                 'cost': z.standard_price
        #                             }
        #                             li.append((0, 0, values))

        return ""

    # Discarded function
    def set_products(self):
        for transaction in self:
            if transaction.transaction_line_id:
                li = []
                for t_line in transaction.transaction_line_id:
                    if t_line.detail_product_category_id and t_line.usr_input:
                        prod_temp_ids = self.env["product.template"].search(
                            [
                                ("option", "=", t_line.usr_input),
                                ("categ_id", "=", t_line.detail_product_category_id.id),
                            ]
                        )
                        for j in prod_temp_ids:
                            prod_ids = self.env["product.product"].search(
                                [("product_tmpl_id", "=", j.id)]
                            )
                            if prod_ids:
                                for i in prod_ids:
                                    values = {
                                        "product_id": i.id,
                                        "cost": i.standard_price,
                                    }
                                    transaction.qbom_line_id.unit_price = values["cost"]
                                    li.append((0, 0, values))
                        if not prod_temp_ids:
                            prod_prod_ids = self.env["product.product"].search(
                                [
                                    ("option", "=", t_line.usr_input),
                                    ("categ_id", "=",
                                     t_line.detail_product_category_id.id),
                                ]
                            )
                            if prod_prod_ids:
                                for z in prod_prod_ids:
                                    values = {
                                        "product_id": z.id,
                                        "cost": z.standard_price,
                                    }
                                    li.append((0, 0, values))

        return li

    # Discarded function
    @api.multi
    def get_calulation(self):
        for transaction in self:
            if transaction.transaction_line_id:
                if transaction.excel_formula:
                    file_con = BytesIO(
                        base64.b64decode(transaction.excel_formula))
                else:
                    raise UserError(_("Excel Formula doesnot Exists"))
                # wb = openpyxl.load_workbook(filename=file_con)
                wb1 = openpyxl.load_workbook(filename=file_con)
                sh = wb1.get_sheet_by_name(transaction.excel_sheet_name)
                wb2 = xlrd.open_workbook(file_contents=base64.b64decode(
                    codecs.decode(transaction.excel_formula)))
                # Read Output
                for t_line2 in transaction.transaction_line_id:
                    sh1 = wb2.sheet_by_name(transaction.excel_sheet_name)
                    row = sh[t_line2.trans_parameter].row-1
                    column = sh[t_line2.trans_parameter].column-1
                    print('XLSX File Output=======>>>>',
                          row, column, sh1.cell(8, 12).value)
                    print('XLSX File Output=======>>>>',
                          row, column, sh1.cell(8, 9).value)
                    t_line2.usr_input = sh1.cell(row, column).value
        self.product_ids = [j for j in self.set_products()]

    # Create method
    @api.model
    def create(self, vals):
        if vals:
            vals['name'] = self.env['ir.sequence'].next_by_code(
                "transaction.info")
        new_id = super(TransactionInfo, self).create(vals)
        return new_id


class TransactionDetail(models.Model):
    _name = 'transaction.details'

    transaction_id = fields.Many2one("transaction.info")
    trans_name = fields.Char("Name", store=True)
    trans_parameter = fields.Char("Excel Cell Value", store=True)
    trans_uom = fields.Char("UOM", store=True)
    usr_input = fields.Char("User Input")
    trans_formula_expression = fields.Char("Formula Expression", store=True)
    detail_product_category_id = fields.Many2one(
        "product.category", string="Product Category")
    check_button = fields.Boolean("Check box")
    trans_process = fields.Selection(
        string="Process", selection=[("Input", "Input"), ("Output", "Output")]
    )


class TransactionDetailProduct(models.Model):
    _name = 'transaction.details.products'

    transaction_id = fields.Many2one("transaction.info")
    product_id = fields.Many2one("product.product")
    cost = fields.Float("Cost")
    description = fields.Char(string="Description", required=False)
    weight = fields.Float(string="Weight")
    product_qty = fields.Float(string="Quantity")
    product_uom_id = fields.Many2one('product.uom', string="UOM")
    price_subtotal = fields.Float(
        string="Subtotal", compute="set_price_subtotal")
    functional_categ_id = fields.Many2one(
        'product.category', string="Functional Category")

    @api.depends('cost')
    def set_price_subtotal(self):
        for order in self:
            order.update({
                'price_subtotal': (order.product_qty * order.cost)
            })

    @api.onchange('product_id')
    def set_product_desc(self):
        for order in self:
            if order.product_id and order.description != '':
                order.description = order.product_id.name
                order.weight = order.product_id.weight
                order.cost = order.product_id.uom_id._compute_price(
                    order.product_id.lst_price, order.product_uom_id)

    @api.model
    def create(self, vals):
        print("vals", vals)
        new_id = super(TransactionDetailProduct, self).create(vals)
        values = {
            "name": new_id.description,
            "product_id": new_id.product_id.id,
            "product_tmpl_id": new_id.product_id.product_tmpl_id.id,
            "product_qty": new_id.product_qty,
            "product_uom_id": new_id.product_uom_id.id,
            "product_category_id": new_id.product_id.categ_id.id,
            "purchase_price": new_id.product_id.standard_price,
            "new_price_unit": new_id.product_id.lst_price,
            "level": '1',
            # 'immi_parent_id': new_id.transaction_id.qbom_line_id.product_id.id,
            'bom_parent_id': new_id.transaction_id.qbom_line_id.id,
            'qbom_id': new_id.transaction_id.qbom_line_id.qbom_id.id,
            'functional_categ_id': new_id.functional_categ_id.id
        }
        qbom_id = self.env['quotation.bom.line']
        qbom_id.create(values)
